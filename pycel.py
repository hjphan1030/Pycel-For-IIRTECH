import os
import glob
import openpyxl
from openpyxl.styles import Font
import re
import tkinter as tk
from tkinter import ttk

# Create the input subfolder if it doesn't exist
if not os.path.exists('input'):
    os.makedirs('input')

# Find all .xlsx files in the input folder
files = glob.glob('./input/*.xlsx')

if not files:
    print("No .xlsx file found in the 'input' folder!")
    exit()

# Create the output subfolder if it doesn't exist
if not os.path.exists('output'):
    os.makedirs('output')

def excel_to_regex(pattern):
    # Escape special regex characters
    pattern = re.escape(pattern)
    # Replace Excel wildcards with regex equivalents
    pattern = pattern.replace("\\*", ".*?").replace("\\?", ".")
    return pattern

# Initialize a dictionary for the replacements
db_replacements = {}

# Check if the db.txt file exists
if os.path.exists('db.txt'):
    # Load the replacements from the db.txt file
    with open('db.txt', 'r', encoding='utf-8') as f:
        db_lines = f.read().splitlines()
        db_replacements = dict(zip(db_lines[::2], db_lines[1::2]))

# Check if the db.xlsx file exists
elif os.path.exists('db.xlsx'):
    # Load workbook
    wb = openpyxl.load_workbook('db.xlsx')
    ws = wb.active
    # Load the replacements from the db.xlsx file
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        old, new = row
        db_replacements[old] = new

# If neither db.txt nor db.xlsx exists, exit the program
else:
    print("No 'db.txt' or 'db.xlsx' file found!")
    exit()

# Create a GUI window
root = tk.Tk()
root.geometry('500x300')
root.title("Pycel processing files...")

# Create the file name text
current_file_name = tk.StringVar()
file_name_label = tk.Label(root, textvariable=current_file_name, font=("", 14))
file_name_label.place(relx=0.5, rely=0.3, anchor='center')

# Create the progress bar
progress = ttk.Progressbar(root, length=200, mode='determinate')
progress.place(relx=0.5, rely=0.5, anchor='center')

# Create the progress text
progress_text = tk.StringVar()
progress_label = tk.Label(root, textvariable=progress_text, font=("", 14))
progress_label.place(relx=0.5, rely=0.6, anchor='center')

# Iterate over all .xlsx files found
for file_index, filename in enumerate(files):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet
    sheet = wb.active

    # Iterate over the cells and replace strings including wildcards
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for old, new in db_replacements.items():
                    # Skip if old is None
                    if old is None:
                        continue
                    # Convert Excel wildcard to regex
                    old = excel_to_regex(old)
                    cell.value = re.sub(old, '' if new is None else str(new), cell.value)

    # Initialize an empty string for the sentence and a list for all sentences
    sentence = ''
    sentences = []
    index = []
    indices = []
    save_sentence_index = 0

    # Iterate over the cells in column 'I' starting from the second row
    for sentence_index, cell in enumerate(sheet['I'][1:], start=1):
        if cell.value and isinstance(cell.value, str):
            save_sentence_index = sentence_index
            # record the starting index of each sentence
            if not index:
                index.append(sheet['A'][sentence_index].value)
                index.append(sheet['B'][sentence_index].value)
                index.append(sheet['C'][sentence_index].value)
                index.append(sheet['D'][sentence_index].value)

            # Add the cell value to the sentence
            sentence += cell.value

            # If the cell value ends with '.' or '?', add the sentence to the list and start a new sentence
            if cell.value.endswith('.') or cell.value.endswith('?'):
                sentences.append(sentence)
                sentence = ''

                index.append(sheet['E'][sentence_index].value)
                index.append(sheet['F'][sentence_index].value)
                index.append(sheet['G'][sentence_index].value)
                index.append(sheet['H'][sentence_index].value)
                indices.append(index)
                index = []
            else:
                sentence += ' '
    if len(sentence) > 0:
        sentences.append(sentence)
        index.append(sheet['E'][save_sentence_index].value)
        index.append(sheet['F'][save_sentence_index].value)
        index.append(sheet['G'][save_sentence_index].value)
        index.append(sheet['H'][save_sentence_index].value)
        indices.append(index)
        index = []

    # Create a new workbook for the sentences
    wb_sentences = openpyxl.Workbook()
    sheet_sentences = wb_sentences.active

    # Write the sentences to the new workbook
    sheet_sentences.cell(row=1, column=1, value='start_h')
    sheet_sentences.cell(row=1, column=2, value='start_m')
    sheet_sentences.cell(row=1, column=3, value='start_s')
    sheet_sentences.cell(row=1, column=4, value='start_ms')
    sheet_sentences.cell(row=1, column=5, value='end_h')
    sheet_sentences.cell(row=1, column=6, value='end_m')
    sheet_sentences.cell(row=1, column=7, value='end_s')
    sheet_sentences.cell(row=1, column=8, value='end_ms')
    sheet_sentences.cell(row=1, column=9, value='내용')
    for sentence_index, sentence in enumerate(sentences, start=1):
        sheet_sentences.cell(row=sentence_index + 1, column=1, value=indices[sentence_index-1][0])
        sheet_sentences.cell(row=sentence_index + 1, column=2, value=indices[sentence_index-1][1])
        sheet_sentences.cell(row=sentence_index + 1, column=3, value=indices[sentence_index-1][2])
        sheet_sentences.cell(row=sentence_index + 1, column=4, value=indices[sentence_index-1][3] + 1)
        sheet_sentences.cell(row=sentence_index + 1, column=5, value=indices[sentence_index-1][4])
        sheet_sentences.cell(row=sentence_index + 1, column=6, value=indices[sentence_index-1][5])
        sheet_sentences.cell(row=sentence_index + 1, column=7, value=indices[sentence_index-1][6])
        sheet_sentences.cell(row=sentence_index + 1, column=8, value=indices[sentence_index-1][7])
        sheet_sentences.cell(row=sentence_index + 1, column=9, value=sentence)

    # Create a font
    font = Font(size=12)

    # Set the height of all rows
    for row in sheet_sentences.iter_rows():
        sheet_sentences.row_dimensions[row[0].row].height = 18

    # Set the width of all columns
    for column in sheet_sentences.columns:
        sheet_sentences.column_dimensions[column[0].column_letter].width = 10

    # Set the font of all cells
    for row in sheet_sentences.iter_rows():
        for cell in row:
            cell.font = font

    # Save the workbook to a new file in the subfolder
    # Use the original filename, but replace the directory
    new_filename_sentences = './output/' + os.path.basename(filename)
    wb_sentences.save(new_filename_sentences)

    # Update the progress bar and the progress text
    progress['value'] = (file_index + 1) / len(files) * 100
    current_file_name.set(os.path.basename(filename))
    if file_index + 1 == len(files):
        progress_text.set(f"{file_index + 1} / {len(files)} files complete! 😄")
    else:
        progress_text.set(f"{file_index + 1} / {len(files)} files")
    root.update()

# Delay the closing of the GUI window by 2.5 seconds
root.after(2000, root.destroy)
root.mainloop()